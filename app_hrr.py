import streamlit as st
from openpyxl import Workbook, load_workbook
import pandas as pd
import os
import matplotlib.pyplot as plt
import numpy as np
from scipy.stats import norm

st.set_page_config(page_title="Registro HRR", layout="centered")

# 🧠 Clasificación
def clasificar_discromatopsia(respuestas):
    rojo_verde = sum([r == '✗' for r in respuestas[4:10]])  # Láminas 5–10
    azul_amarillo = 0
    for i in range(10, len(respuestas), 2):  # Láminas 11–24 (Normal + Defecto)
        if respuestas[i] == '✗' or respuestas[i+1] == '✗':
            azul_amarillo += 1

    total_errores = rojo_verde + azul_amarillo

    if rojo_verde >= 3:
        tipo = 'Protan/Deutan'
    elif azul_amarillo >= 2:
        tipo = 'Tritan'
    elif total_errores >= 5:
        tipo = 'Combinado'
    else:
        tipo = 'Normal'

    if total_errores == 0:
        severidad = '—'
    elif total_errores <= 2:
        severidad = 'Leve'
    elif total_errores <= 4:
        severidad = 'Moderada'
    else:
        severidad = 'Severa'

    return tipo, severidad

# 📁 Guardar en Excel
def guardar_excel(fila):
    archivo = "resultados_HRR.xlsx"
    if not os.path.exists(archivo):
        wb = Workbook()
        ws = wb.active
        encabezado = ["ID", "Edad"] + \
                     [f"Lámina {i+1}" for i in range(10)] + \
                     [f"Lámina {i+1} (Normal)" for i in range(10, 24)] + \
                     [f"Lámina {i+1} (Defecto)" for i in range(10, 24)] + \
                     ["Tipo", "Severidad"]
        ws.append(encabezado)
    else:
        wb = load_workbook(archivo)
        ws = wb.active

    ws.append(fila)
    wb.save(archivo)

# 📊 Cargar datos
def cargar_datos():
    archivo = "resultados_HRR.xlsx"
    if os.path.exists(archivo):
        return pd.read_excel(archivo)
    return pd.DataFrame()

# 📈 Graficar y guardar
def plot_and_save(series, title, filename):
    fig, ax = plt.subplots()
    series.plot(kind='bar', ax=ax, color='skyblue')
    ax.set_title(title)
    ax.set_ylabel("Frecuencia")
    plt.tight_layout()
    fig.savefig(filename)
    return filename

# 📐 Z-test
def z_test(p_obs, p_exp, n):
    se = np.sqrt(p_exp * (1 - p_exp) / n)
    z = (p_obs - p_exp) / se
    p_value = 2 * (1 - norm.cdf(abs(z)))
    return z, p_value

# 🧾 Registro
st.title("Registro de prueba HRR")

with st.form("registro_form"):
    id_estudiante = st.text_input("ID del estudiante")
    edad = st.number_input("Edad", min_value=0, max_value=120, step=1)
    respuestas = []
    cols = st.columns(6)

    # Láminas 1–10: una sola respuesta
    for i in range(10):
        with cols[i % 6]:
            r = st.selectbox(f"Lámina {i+1}", ["✓", "✗"], key=f"lamina_{i}")
            respuestas.append(r)

    # Láminas 11–24: dos respuestas por lámina
    for i in range(10, 24):
        with cols[i % 6]:
            r1 = st.selectbox(f"Lámina {i+1} (Normal)", ["✓", "✗"], key=f"lamina_{i}_n")
            r2 = st.selectbox(f"Lámina {i+1} (Defecto)", ["✓", "✗"], key=f"lamina_{i}_d")
            respuestas.append(r1)
            respuestas.append(r2)

    submitted = st.form_submit_button("Guardar resultado")

    if submitted:
        tipo, severidad = clasificar_discromatopsia(respuestas)
        fila = [id_estudiante, edad] + respuestas + [tipo, severidad]
        guardar_excel(fila)
        st.success(f"Resultado guardado: {tipo} ({severidad})")

# 📊 Estadísticas
st.header("Estadísticas acumuladas")
df = cargar_datos()

if not df.empty:
    st.sidebar.header("🔍 Filtros")
    edad_min = st.sidebar.slider("Edad mínima", min_value=0, max_value=120, value=18)
    edad_max = st.sidebar.slider("Edad máxima", min_value=0, max_value=120, value=30)
    df_filtrado = df[(df["Edad"] >= edad_min) & (df["Edad"] <= edad_max)]

    st.write(f"Total de registros: {len(df_filtrado)}")

    col1, col2 = st.columns(2)
    with col1:
        st.subheader("Distribución por tipo")
        tipo_counts = df_filtrado["Tipo"].value_counts()
        st.bar_chart(tipo_counts)

    with col2:
        st.subheader("Distribución por severidad")
        sev_counts = df_filtrado["Severidad"].value_counts()
        st.bar_chart(sev_counts)

    tipo_img = plot_and_save(tipo_counts, "Distribución por tipo", "tipo.png")
    sev_img = plot_and_save(sev_counts, "Distribución por severidad", "severidad.png")

    with st.expander("📸 Descargar gráficos"):
        with open(tipo_img, "rb") as f:
            st.download_button("Descargar gráfico de tipo", f, file_name="tipo.png")
        with open(sev_img, "rb") as f:
            st.download_button("Descargar gráfico de severidad", f, file_name="severidad.png")

    st.subheader("📐 Prueba de hipótesis")
    casos = df_filtrado[df_filtrado["Tipo"] != "Normal"]
    p_obs = len(casos) / len(df_filtrado)
    p_exp = 0.076
    z, p = z_test(p_obs, p_exp, len(df_filtrado))

    st.write(f"Prevalencia observada: {p_obs:.3f}")
    st.write(f"Prevalencia esperada: {p_exp:.3f}")
    st.write(f"Z = {z:.2f}, p = {p:.4f}")

    if p < 0.05:
        st.error("La diferencia es estadísticamente significativa (p < 0.05)")
    else:
        st.success("No hay diferencia significativa (p ≥ 0.05)")

    st.download_button("📥 Descargar Excel", data=df_filtrado.to_excel(index=False), file_name="resultados_HRR.xlsx")
else:
    st.info("No hay datos registrados aún.")

